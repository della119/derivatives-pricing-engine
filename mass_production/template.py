"""
Excel input template generator for batch (mass production) valuations.

Layout
------
Sheet 1 — "Instructions"     : Plain-language usage guide
Sheet 2 — "Product Reference": Lookup table of all 47 products with their
                                expected parameters and option-type rules
Sheet 3 — "Valuation Input"  : The user-fillable input table
                                (one row per trade)

The system reads only Sheet 3 ("Valuation Input"). Sheets 1 and 2 are for
human reference.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from product_registry import PRODUCTS, CATEGORIES, get_category_name


# ===================================================================
# Styling helpers
# ===================================================================
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="1F4E78", size=11)
NOTE_FONT = Font(name="Calibri", italic=True, color="555555", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ===================================================================
# Standard input columns (the system uses these column names exactly)
# ===================================================================
INPUT_COLUMNS = [
    "Trade ID",
    "Product Key",
    "Option Type",
    "Notional",
    "S",        # Spot
    "K",        # Strike
    "T",        # Time to expiry (years)
    "r",        # Risk-free rate
    "b",        # Cost of carry (optional)
    "q",        # Dividend yield (optional)
    "sigma",    # Volatility
    "Extra Params",  # JSON-style for less-common fields, optional
    "Notes",
]

INPUT_COLUMN_HELP = {
    "Trade ID":      "Free-text identifier (e.g., TRD001, BookA-001).",
    "Product Key":   "Product key from the 'Product Reference' sheet (e.g., black_scholes).",
    "Option Type":   "call / put. Leave blank if product has no option type.",
    "Notional":      "Position size (number of contracts/units). Default 1.",
    "S":             "Underlying spot price.",
    "K":             "Strike price.",
    "T":             "Time-to-expiry in YEARS (e.g., 0.5 for 6-month).",
    "r":             "Annualised risk-free rate as a decimal (e.g., 0.05 = 5%).",
    "b":             "Cost-of-carry (optional, defaults to r if blank).",
    "q":             "Continuous dividend yield as decimal (optional).",
    "sigma":         "Annualised volatility as decimal (e.g., 0.20 = 20%).",
    "Extra Params":  "Optional JSON for advanced products, e.g., {\"H\":110, \"barrier_type\":\"up_in_call\"}.",
    "Notes":         "Free-text comments — ignored by the system.",
}


# ===================================================================
# Sheet builders
# ===================================================================
def _build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 110

    rows = [
        ("Derivatives Pricing Engine — Mass Production Template", "title"),
        ("", None),
        ("How to use this template", "h1"),
        ("1. Open the 'Product Reference' sheet to find the Product Key for each product you want to price.", "p"),
        ("2. Open the 'Valuation Input' sheet and fill in one row per trade.", "p"),
        ("3. Save the file and upload it back to the Pricing Engine sidebar — the system will price everything in batch and return an Excel results file.", "p"),
        ("", None),
        ("Required columns", "h1"),
        ("• Product Key   — must match a key in the 'Product Reference' sheet.", "p"),
        ("• S, K, T, r, sigma  — most products require these five core parameters.", "p"),
        ("• Option Type   — fill in 'call' or 'put' for products that have one (see Product Reference).", "p"),
        ("", None),
        ("Optional columns", "h1"),
        ("• Notional      — number of contracts. Defaults to 1 if blank.", "p"),
        ("• b, q          — cost-of-carry, dividend yield. Defaults to r and 0.", "p"),
        ("• Extra Params  — JSON for less-common fields (e.g., barrier level, jump intensity).", "p"),
        ("                  Example: {\"H\": 110, \"barrier_type\": \"up_in_call\"}", "code"),
        ("• Notes         — free-text, ignored by the engine.", "p"),
        ("", None),
        ("Output", "h1"),
        ("After processing, you will receive an Excel file with these columns added per row:", "p"),
        ("    Price, Delta, Gamma, Vega, Theta, Rho, Status, Error", "code"),
        ("", None),
        ("Conventions", "h1"),
        ("• Rates and yields are decimals (0.05 = 5%, NOT 5).", "p"),
        ("• Volatilities are decimals (0.20 = 20%).", "p"),
        ("• Time is in years (0.25 = 3 months).", "p"),
        ("• Position direction is implied long for valuation; multiply by -1 in your own books for shorts.", "p"),
        ("", None),
        ("Tips", "h1"),
        ("• Start with a few rows to validate, then scale up.", "p"),
        ("• If a row errors out, check the 'Error' column in the output for the reason.", "p"),
        ("• Exotic products may need 'Extra Params'; refer to the 'Product Reference' sheet for hints.", "p"),
    ]

    r = 1
    for text, kind in rows:
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E78")
        elif kind == "h1":
            cell.font = Font(name="Calibri", bold=True, size=12, color="1F4E78")
            cell.fill = SUBHEAD_FILL
        elif kind == "code":
            cell.font = Font(name="Consolas", size=10, color="333333")
        else:
            cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22 if kind == "title" else (18 if kind == "h1" else 16)
        r += 1


def _build_product_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Product Reference", 1)

    headers = ["Product Key", "Category", "Display Name", "Has Option Type",
               "Required Params", "Description"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

    # Order products by category
    cat_order = {key: i for i, (key, _) in enumerate(CATEGORIES)}
    ordered = sorted(
        PRODUCTS.items(),
        key=lambda kv: (cat_order.get(kv[1].get("category", ""), 99), kv[0]),
    )

    row = 2
    for pid, p in ordered:
        params_str = ", ".join(par["key"] for par in p.get("params", []))
        ws.cell(row=row, column=1, value=pid).font = Font(name="Consolas", size=10, bold=True)
        ws.cell(row=row, column=2, value=get_category_name(p.get("category", "")))
        ws.cell(row=row, column=3, value=p.get("name", ""))
        ws.cell(row=row, column=4, value="Yes" if p.get("has_option_type") else "No")
        ws.cell(row=row, column=5, value=params_str).font = Font(name="Consolas", size=10)
        ws.cell(row=row, column=6, value=p.get("desc", ""))
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    widths = [22, 28, 38, 14, 36, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_valuation_input_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Valuation Input", 2)

    # Header row
    for col, h in enumerate(INPUT_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

    # Help row (italic, light grey, in row 2)
    for col, h in enumerate(INPUT_COLUMNS, 1):
        c = ws.cell(row=2, column=col, value=INPUT_COLUMN_HELP.get(h, ""))
        c.font = NOTE_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Two example rows
    examples = [
        ["TRD001", "black_scholes", "call", 100, 100, 100, 1.0, 0.05, "", "", 0.20, "", "ATM 1Y call"],
        ["TRD002", "black_scholes", "put",  100, 100,  95, 0.5, 0.05, "", "", 0.25, "", "OTM 6M put"],
        ["TRD003", "merton_1973",  "call", 50,  120, 110, 2.0, 0.04, "", 0.02, 0.18, "", "Dividend-paying stock"],
    ]
    for r_idx, row_vals in enumerate(examples, start=3):
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.alignment = Alignment(vertical="center")
            c.border = THIN_BORDER
            if col == 2:  # Product Key
                c.font = Font(name="Consolas", size=11)

    widths = [12, 24, 13, 11, 9, 9, 9, 9, 9, 9, 9, 30, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 60
    ws.freeze_panes = "A3"


# ===================================================================
# Public API
# ===================================================================
def build_template_workbook() -> Workbook:
    """Return a freshly-built openpyxl Workbook (3 sheets)."""
    wb = Workbook()
    # remove the default empty sheet
    default = wb.active
    wb.remove(default)

    _build_instructions_sheet(wb)
    _build_product_reference_sheet(wb)
    _build_valuation_input_sheet(wb)

    return wb


def get_template_bytes() -> bytes:
    """Build the template and return it as bytes (for st.download_button)."""
    wb = build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
