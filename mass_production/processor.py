"""
Process an uploaded Excel template — run valuations and Greeks for every
row in the 'Valuation Input' sheet, then write a results workbook.
"""
from __future__ import annotations

import io
import json
import importlib
from typing import Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from product_registry import PRODUCTS


# ===================================================================
# Caching for pricer imports
# ===================================================================
_PRICER_CACHE: dict = {}


def _load_pricer(module_path: str, func_name: str):
    key = (module_path, func_name)
    if key not in _PRICER_CACHE:
        mod = importlib.import_module(module_path)
        _PRICER_CACHE[key] = getattr(mod, func_name)
    return _PRICER_CACHE[key]


# ===================================================================
# Greeks (lightweight bump-and-reprice — re-uses logic similar to app.py)
# ===================================================================
def _numerical_greeks(pricer, kwargs, S_key="S", sigma_key="sigma",
                      T_key="T", r_key="r"):
    """Compute Δ, Γ, Vega, Theta, Rho via finite differences."""
    greeks = {"Delta": np.nan, "Gamma": np.nan, "Vega": np.nan,
              "Theta": np.nan, "Rho": np.nan}
    try:
        base = float(pricer(**kwargs))
    except Exception:
        return greeks, np.nan

    # Delta & Gamma
    if S_key in kwargs:
        S = float(kwargs[S_key])
        dS = max(S * 0.001, 0.01)
        try:
            up = float(pricer(**{**kwargs, S_key: S + dS}))
            dn = float(pricer(**{**kwargs, S_key: S - dS}))
            greeks["Delta"] = (up - dn) / (2 * dS)
            greeks["Gamma"] = (up - 2 * base + dn) / (dS * dS)
        except Exception:
            pass

    # Vega (per 1% vol)
    if sigma_key in kwargs:
        sig = float(kwargs[sigma_key])
        d = 0.001
        try:
            up = float(pricer(**{**kwargs, sigma_key: sig + d}))
            dn = float(pricer(**{**kwargs, sigma_key: max(sig - d, 1e-4)}))
            greeks["Vega"] = (up - dn) / (2 * d) * 0.01
        except Exception:
            pass

    # Theta (per day)
    if T_key in kwargs and float(kwargs[T_key]) > 1 / 365:
        T = float(kwargs[T_key])
        try:
            short = float(pricer(**{**kwargs, T_key: T - 1 / 365}))
            greeks["Theta"] = -(base - short)
        except Exception:
            pass

    # Rho (per 1% rate)
    if r_key in kwargs:
        r = float(kwargs[r_key])
        d = 0.0001
        try:
            up = float(pricer(**{**kwargs, r_key: r + d}))
            dn = float(pricer(**{**kwargs, r_key: r - d}))
            greeks["Rho"] = (up - dn) / (2 * d) * 0.01
        except Exception:
            pass

    return greeks, base


# ===================================================================
# Row processing
# ===================================================================
def _build_kwargs_for_product(product: dict, row: pd.Series) -> Tuple[dict, str]:
    """
    Build the kwargs dict to pass to the pricing function for this row.
    Only keys that the product expects are included.
    Returns (kwargs, error_message). If error_message is non-empty, the row
    failed validation and should not be priced.
    """
    expected_keys = {p["key"] for p in product.get("params", [])}
    kwargs: dict = {}

    # Helper: pull numeric, blank-tolerant
    def _num(col_name):
        if col_name not in row:
            return None
        v = row[col_name]
        if v is None or (isinstance(v, float) and np.isnan(v)) or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # Map standard columns
    standard = {"S": "S", "K": "K", "T": "T", "r": "r", "b": "b",
                "q": "q", "sigma": "sigma"}
    for col, key in standard.items():
        if key in expected_keys:
            v = _num(col)
            if v is not None:
                kwargs[key] = v

    # Cost-of-carry default: b = r if expected but blank
    if "b" in expected_keys and "b" not in kwargs and "r" in kwargs:
        kwargs["b"] = kwargs["r"]

    # Option type
    if product.get("has_option_type"):
        ot = row.get("Option Type", "")
        if isinstance(ot, str):
            ot = ot.strip().lower()
        if ot not in ("call", "put") and product.get("option_type_choices"):
            choices = product["option_type_choices"]
            if ot in choices:
                pass
            else:
                return {}, f"Option Type must be one of {choices}"
        kwargs["option_type"] = ot or "call"

    # Extra Params (JSON)
    extra_raw = row.get("Extra Params", "")
    if isinstance(extra_raw, str) and extra_raw.strip():
        try:
            extra = json.loads(extra_raw)
            if isinstance(extra, dict):
                kwargs.update(extra)
        except json.JSONDecodeError as e:
            return {}, f"Extra Params JSON parse error: {e}"

    # Validate required keys are present
    missing = [k for k in expected_keys
               if k not in kwargs and k not in ("b",)]
    if missing:
        return {}, f"Missing required parameter(s): {', '.join(missing)}"

    # Cast integer-like params
    int_keys = {"N", "N_S", "N_T", "n", "n_paths", "n_steps", "n_terms"}
    for k in int_keys & set(kwargs.keys()):
        try:
            kwargs[k] = int(kwargs[k])
        except (TypeError, ValueError):
            pass

    return kwargs, ""


def _process_row(row: pd.Series) -> dict:
    """Run a single valuation row. Returns a dict of output columns."""
    out = {
        "Price": np.nan, "Delta": np.nan, "Gamma": np.nan,
        "Vega": np.nan, "Theta": np.nan, "Rho": np.nan,
        "Status": "OK", "Error": "",
    }

    pkey = str(row.get("Product Key", "")).strip()
    if not pkey:
        out["Status"] = "Error"
        out["Error"] = "Product Key is empty"
        return out

    if pkey not in PRODUCTS:
        out["Status"] = "Error"
        out["Error"] = f"Unknown Product Key '{pkey}' (see Product Reference sheet)"
        return out

    product = PRODUCTS[pkey]

    # Skip Monte Carlo products in batch mode (they return dicts and are slow)
    if product.get("returns_dict"):
        out["Status"] = "Skipped"
        out["Error"] = "Monte Carlo products are not supported in batch mode"
        return out

    kwargs, err = _build_kwargs_for_product(product, row)
    if err:
        out["Status"] = "Error"
        out["Error"] = err
        return out

    # Notional (defaults to 1)
    try:
        notional = float(row.get("Notional", 1) or 1)
    except (TypeError, ValueError):
        notional = 1.0

    try:
        pricer = _load_pricer(product["module"], product["func"])
        greeks, price = _numerical_greeks(pricer, kwargs)
    except Exception as e:
        out["Status"] = "Error"
        out["Error"] = f"Pricing failed: {e}"
        return out

    if price is None or (isinstance(price, float) and np.isnan(price)):
        out["Status"] = "Error"
        out["Error"] = "Pricer returned NaN/None"
        return out

    out["Price"] = float(price) * notional
    for g in ("Delta", "Gamma", "Vega", "Theta", "Rho"):
        v = greeks.get(g, np.nan)
        out[g] = float(v) * notional if not (isinstance(v, float) and np.isnan(v)) else np.nan
    return out


# ===================================================================
# Workbook I/O
# ===================================================================
RESULT_FILL_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RESULT_FILL_ERR = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RESULT_FILL_SKIP = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def process_uploaded_workbook(file_bytes: bytes) -> Tuple[bytes, dict]:
    """
    Read uploaded workbook, run all valuations, return results workbook bytes
    plus a small summary dict for the UI.
    """
    # Read input — start at row 1 (header), skip help row (row 2)
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name="Valuation Input",
        header=0,
        skiprows=[1],   # skip the italic help row
    )
    # Drop blank rows (no Product Key)
    df = df.dropna(subset=["Product Key"]).reset_index(drop=True)

    if len(df) == 0:
        raise ValueError(
            "No data rows found in 'Valuation Input' sheet. "
            "Did you fill in any rows below the header?"
        )

    # Process each row
    results = [_process_row(row) for _, row in df.iterrows()]
    res_df = pd.DataFrame(results)

    # Combine input + results
    out_df = pd.concat([df.reset_index(drop=True), res_df], axis=1)

    # Build output workbook
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Results", 0)

    headers = list(out_df.columns)
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

    for r_idx, row in enumerate(out_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if not (isinstance(val, float) and np.isnan(val)) else None)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # Format numerics
            col_name = headers[c_idx - 1]
            if col_name == "Price":
                cell.number_format = "#,##0.0000"
            elif col_name in ("Delta", "Gamma", "Vega", "Theta", "Rho"):
                cell.number_format = "#,##0.0000"

        # Status colouring on the whole row
        status = out_df.iloc[r_idx - 2]["Status"]
        fill = RESULT_FILL_OK if status == "OK" else (
            RESULT_FILL_SKIP if status == "Skipped" else RESULT_FILL_ERR
        )
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = fill

    # Column widths
    widths = []
    for h in headers:
        if h in ("Notes", "Error", "Extra Params"):
            widths.append(36)
        elif h in ("Description",):
            widths.append(40)
        elif h == "Product Key":
            widths.append(22)
        else:
            widths.append(13)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Build summary sheet
    sws = wb.create_sheet("Summary", 1)
    n_total = len(out_df)
    n_ok = int((out_df["Status"] == "OK").sum())
    n_err = int((out_df["Status"] == "Error").sum())
    n_skip = int((out_df["Status"] == "Skipped").sum())
    total_value = float(out_df.loc[out_df["Status"] == "OK", "Price"].sum()) if n_ok > 0 else 0.0

    summary_rows = [
        ("Total Rows",   n_total),
        ("Successful",   n_ok),
        ("Errors",       n_err),
        ("Skipped",      n_skip),
        ("Sum of Prices (OK rows)", total_value),
    ]
    for r, (label, val) in enumerate(summary_rows, 1):
        a = sws.cell(row=r, column=1, value=label)
        b = sws.cell(row=r, column=2, value=val)
        a.font = Font(name="Calibri", bold=True, size=11)
        b.font = Font(name="Calibri", size=11)
        if isinstance(val, float):
            b.number_format = "#,##0.0000"
    sws.column_dimensions["A"].width = 30
    sws.column_dimensions["B"].width = 18

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    summary = {
        "total": n_total,
        "ok": n_ok,
        "errors": n_err,
        "skipped": n_skip,
        "total_value": total_value,
    }
    return buf.getvalue(), summary
